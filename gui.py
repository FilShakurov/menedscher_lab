import sys
import os
import traceback
from datetime import datetime
from pathlib import Path


from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget,
                             QVBoxLayout, QLabel, QPushButton, QHBoxLayout, QLineEdit,
                             QListWidget, QListWidgetItem, QFileDialog, QMessageBox,
                             QDialog, QPlainTextEdit, QStackedWidget)
from PyQt5.QtCore import Qt
import qdarkstyle
from openpyxl import load_workbook
from orkestrator_db import MainCore
from gransostav import RaschetGranov
from config_core.core import zagr_file, zagr_file2, zagr_tarirovki, obrabotka_df_posle_zagr, rashet_gran, vigruzka_namiv
from config_core import config
from config_core.klasspredict import ClassPredict
import gran_sync
from gran_report_dialog import GranReportDialog

GROUP_DIR = Path(r"Y:\2026\Группа физических и механических испытаний")

class NewQDialog(QDialog):
    def __init__(self, db, parent=None):
        super().__init__()
        self.setWindowTitle("Добавить объект")
        self.resize(400, 200)

        self.orkestr_db = db

        main_layout = QVBoxLayout(self)

        main_layout.addWidget(QLabel("Название объекта:"))
        self.lineedit = QLineEdit('')
        main_layout.addWidget(self.lineedit)

        main_layout.addWidget(QLabel("Год:"))
        self.year_edit = QLineEdit(str(datetime.now().year))
        main_layout.addWidget(self.year_edit)

        self.btn = QPushButton("Добавить объект")
        self.btn.clicked.connect(self.btn_calc)
        main_layout.addWidget(self.btn)

    def btn_calc(self):
        try:
            name = self.lineedit.text().strip()
            if not name:
                QMessageBox.warning(self, "Внимание", "Введите название объекта")
                return
            year_str = self.year_edit.text().strip()
            year = int(year_str) if year_str.isdigit() else None
            self.orkestr_db.db_add.add_object_bd(name, year=year)
            QMessageBox.information(self, "Успех", f"Объект '{name}' добавлен.")
        except Exception as e:
            print(e)
            traceback.print_exc()
            QMessageBox.critical(self, "Ошибка", str(e))


class NewQDialog2(QDialog):
    """Добавление новой партии в объект, уже выбранный на шаге 1 мастера."""

    def __init__(self, db, object_id, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Добавить партию")
        self.resize(420, 150)

        self.orkestr_db = db
        self.object_id = object_id

        main_layout = QVBoxLayout(self)

        main_layout.addWidget(QLabel("Выберите файл рабочей сводной для новой партии:"))

        self.btn = QPushButton("Выбрать файл и добавить партию")
        self.btn.clicked.connect(self.btn_calc)
        main_layout.addWidget(self.btn)

    def btn_calc(self):
        try:
            path_rab_svodn, _ = QFileDialog.getOpenFileName(
                self, "Выберите файл рабочей сводной", "", "Excel Files (*.xlsx *.xls)")
            if not path_rab_svodn:
                return

            # Получаем только имя файла
            file_name = os.path.basename(path_rab_svodn)

            partiya_id = self.orkestr_db.db_add.add_partiya_bd(file_name, self.object_id)

            self.add_rab_svodn(path_rab_svodn, partiya_id)

            self.accept()

        except Exception as e:
            print(e)
            traceback.print_exc()

    def add_rab_svodn(self, path_rab_svodn, part_id):

        db_id = part_id

        try:
            df = zagr_file2(path_rab_svodn)
            self.orkestr_db.db_add.add_rab_svodnaya(df, db_id)

            df2 = RaschetGranov.zagr_excel(path_rab_svodn)
            df2 = RaschetGranov.raschet_gran_pesk(df2)
            self.orkestr_db.db_add.add_gran_bd(df2)

            # --- Сохраняем путь к файлу и дату его изменения в БД ---
            mtime_str = gran_sync.get_file_mtime_str(path_rab_svodn)

            # Также парсим статусы грансоставов (монолит/нарушен) из файла
            samples = gran_sync.parse_rab_svodn_excel(path_rab_svodn)
            monoliths, disturbed = gran_sync.count_sample_types(samples)

            self.orkestr_db.db.update_partii_file_info(
                db_id, path_rab_svodn, mtime_str, monoliths, disturbed
            )

            # Записываем sample_type и начальные статусы грансоставов
            if samples:
                self.orkestr_db.db.update_probi_sample_type_and_status(db_id, samples)

            QMessageBox.information(
                self, "Успех",
                f"Рабочая сводная загружена.\n"
                f"Монолиты: {monoliths} шт. | Нарушены: {disturbed} шт.\n"
                f"Путь к файлу сохранён в БД."
            )
        except Exception as e:
            print(e)
            traceback.print_exc()
            QMessageBox.critical(self, "Ошибка", f"Не удалось загрузить рабочую сводную: {e}")


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        self.orkestr_db = MainCore('database.db')

        self.df_partii = None
        self.current_object_id = None
        self.current_object_name = None

        self.initUI()

    def initUI(self):
        self.setGeometry(100, 100, 900, 650)
        self.setWindowTitle('Шаблон')

        self.stack = QStackedWidget()
        self.setCentralWidget(self.stack)

        self.stack.addWidget(self._build_object_page())    # index 0 — шаг 1
        self.stack.addWidget(self._build_partiya_page())   # index 1 — шаг 2

        self.reload_objects()
        self.stack.setCurrentIndex(0)

    # ------------------------------------------------------------------
    # Шаг 1 — выбор объекта
    # ------------------------------------------------------------------
    def _build_object_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)

        title = QLabel("Шаг 1 — выберите объект")
        title.setStyleSheet("font-weight: bold; font-size: 14px;")
        layout.addWidget(title)

        objects_hlayout = QHBoxLayout()
        layout.addLayout(objects_hlayout)

        self.list_widget1 = QListWidget()
        self.list_widget1.itemDoubleClicked.connect(self._on_object_chosen)
        self.list_widget1.itemClicked.connect(self._on_object_summary)
        objects_hlayout.addWidget(self.list_widget1)

        self.object_summary_box = QPlainTextEdit()
        self.object_summary_box.setReadOnly(True)
        self.object_summary_box.setPlaceholderText("Выберите объект, чтобы увидеть сводку")
        objects_hlayout.addWidget(self.object_summary_box)

        hlayout = QHBoxLayout()
        layout.addLayout(hlayout)

        self.btn_add_object = QPushButton("Добавить объект")
        self.btn_add_object.clicked.connect(self.add_object)
        hlayout.addWidget(self.btn_add_object)

        self.btn_next = QPushButton("Далее →")
        self.btn_next.clicked.connect(self._go_to_partiya_step)
        hlayout.addWidget(self.btn_next)

        return page

    def _on_object_chosen(self, item):
        self.list_widget1.setCurrentItem(item)
        self._go_to_partiya_step()

    def _on_object_summary(self, item):
        object_id = item.data(Qt.UserRole)
        object_name = item.text()
        try:
            df = self.orkestr_db.db_show.get_probi_by_object(object_id)

            kol_partii = df["partiya_id"].nunique()
            kol_prob = df["proba_id"].nunique()

            summary_lines = [
                f"Объект: {object_name}",
                f"Количество партий: {kol_partii}",
                f"Количество проб: {kol_prob}",
            ]

            # if kol_prob:
            #     by_status = df["status_gran"].fillna("Не назначен").value_counts()
            #     summary_lines.append("")
            #     summary_lines.append("По статусу грансостава:")
            #     for status, count in by_status.items():
            #         summary_lines.append(f"  {status}: {count}")

            self.object_summary_box.setPlainText("\n".join(summary_lines))
        except Exception as e:
            print(e)
            traceback.print_exc()
            self.object_summary_box.setPlainText(f"Не удалось получить сводку: {e}")

    def _go_to_partiya_step(self):
        item = self.list_widget1.currentItem()
        if not item:
            QMessageBox.warning(self, "Внимание", "Выберите объект из списка")
            return

        self.current_object_id = item.data(Qt.UserRole)
        self.current_object_name = item.text()
        self.df_partii = None

        self.lbl_current_object.setText(f"Объект: {self.current_object_name}")
        self.info_label.setText("Отчет:")
        self.warning_box.clear()
        self.reload_partii()

        self.stack.setCurrentIndex(1)

    def reload_objects(self):
        self.list_widget1.clear()
        rows = self.orkestr_db.db_show.show_all_objects()
        for row in rows:
            item = QListWidgetItem(row["name_object"])
            # Сохраняем ID из БД в роли UserRole (или любой другой роли)
            item.setData(Qt.UserRole, row["id"])
            self.list_widget1.addItem(item)

    def add_object(self):
        try:
            dialog = NewQDialog(self.orkestr_db, self)
            dialog.exec_()
            self.reload_objects()
        except Exception as e:
            print(e)
            traceback.print_exc()

    # ------------------------------------------------------------------
    # Шаг 2 — работа с партиями выбранного объекта
    # ------------------------------------------------------------------
    def _build_partiya_page(self):
        page = QWidget()
        main_layout = QVBoxLayout(page)

        header = QHBoxLayout()
        main_layout.addLayout(header)

        self.btn_back = QPushButton("← Назад к объектам")
        self.btn_back.clicked.connect(self._go_back_to_objects)
        header.addWidget(self.btn_back)

        self.lbl_current_object = QLabel("Объект:")
        self.lbl_current_object.setStyleSheet("font-weight: bold; font-size: 14px;")
        header.addWidget(self.lbl_current_object)
        header.addStretch()

        self.btn_namiv = QPushButton("Добавить намыв")
        self.btn_namiv.clicked.connect(self.add_namiv)
        header.addWidget(self.btn_namiv)

        hlayout = QHBoxLayout()
        main_layout.addLayout(hlayout)

        vlayout2 = QVBoxLayout()
        hlayout.addLayout(vlayout2)

        self.list_widget2 = QListWidget()
        self.list_widget2.itemClicked.connect(self.on_item_clicked2)
        self.list_widget2.setFixedWidth(300)
        vlayout2.addWidget(self.list_widget2)

        self.btn2 = QPushButton("Добавить партию")
        self.btn2.clicked.connect(self.add_partiya)
        vlayout2.addWidget(self.btn2)

        vlayout3 = QVBoxLayout()
        hlayout.addLayout(vlayout3)

        self.info_label = QLabel("Отчет:")
        vlayout3.addWidget(self.info_label)

        self.warning_box = QPlainTextEdit()
        vlayout3.addWidget(self.warning_box)

        self.btn6 = QPushButton("Получить все намывы по этой партии")
        self.btn6.clicked.connect(self.get_namivs)
        vlayout3.addWidget(self.btn6)

        self.btn7 = QPushButton("Выгрузить в Excel отчет по партии")
        self.btn7.clicked.connect(self.save_part_excel)
        vlayout3.addWidget(self.btn7)

        self.btn8 = QPushButton("Проверка по статистике")
        self.btn8.clicked.connect(self.btn_calc)
        vlayout3.addWidget(self.btn8)

        # --- Кнопки синхронизации и отчёта грансоставов ---
        self.btn_sync_incr = QPushButton("🔄 Проверить обновления файлов")
        self.btn_sync_incr.setToolTip(
            "Обходит все сохранённые пути к файлам, сравнивает дату изменения.\n"
            "Если файл изменился — добавляет новые данные без потери уже внесённых."
        )
        self.btn_sync_incr.clicked.connect(self.sync_files_incremental)
        vlayout3.addWidget(self.btn_sync_incr)

        self.btn_sync_full = QPushButton("⚠ Перепарсить файлы заново")
        self.btn_sync_full.setToolTip(
            "Полный перепарс всех файлов с перезаписью статусов.\n"
            "Использовать, если нужно полностью обновить данные."
        )
        self.btn_sync_full.clicked.connect(self.sync_files_full)
        vlayout3.addWidget(self.btn_sync_full)

        self.btn_gran_report = QPushButton("📊 Отчёт грансоставов")
        self.btn_gran_report.setToolTip("Открывает окно с фильтрами и статистикой по статусам грансостава")
        self.btn_gran_report.clicked.connect(self.open_gran_report)
        vlayout3.addWidget(self.btn_gran_report)

        return page

    def _go_back_to_objects(self):
        self.current_object_id = None
        self.current_object_name = None
        self.df_partii = None
        self.reload_objects()
        self.stack.setCurrentIndex(0)

    def reload_partii(self):
        self.list_widget2.clear()
        if self.current_object_id is None:
            return
        rows = self.orkestr_db.db_show.show_all_partii_object(self.current_object_id)
        for row in rows:
            item = QListWidgetItem(row["name_partii"])
            item.setData(Qt.UserRole, row["id"])
            self.list_widget2.addItem(item)

    def add_partiya(self):
        if self.current_object_id is None:
            return
        try:
            dialog = NewQDialog2(self.orkestr_db, self.current_object_id, self)
            dialog.exec_()
            self.reload_partii()
        except Exception as e:
            print(e)
            traceback.print_exc()

    def btn_calc(self):
        try:
            df_table = self.df_partii

            bad_rows = ClassPredict.predict(df_table)
            bad_rows = bad_rows.sort_values(by=['lab_nomer'], ascending=True)
            # bad_rows = bad_rows[config.COLUMNS2]

            if not bad_rows.empty:
                messages = []
                for idx, row in bad_rows.iterrows():
                    err_type = row['error_type']
                    if err_type == 'плотность':
                        messages.append(
                            f"{row['lab_nomer']}: {err_type} (испытание, статистика)="
                            f"{row['plotn']:.2f}, "
                            f"{row['plotn_predict']:.2f}"
                        )
                    elif err_type == 'влажность':
                        messages.append(
                            f"{row['lab_nomer']}: {err_type} (испытание, статистика)="
                            f"{row['wlashn']:.3f}, "
                            f"{row['wlashn_predict']:.3f}"
                        )

                    elif err_type == 'Укол':
                        messages.append(
                            f"{row['lab_nomer']}: {err_type} (испытание, статистика)="
                            f"{row['ukol_values']:.1f}, "
                            f"{row['ukol_predict']:.1f}"
                        )

                self.warning_box.setPlainText("\n\n".join(messages))
                QMessageBox.warning(
                    self,
                    "Предупреждение",
                    f"Найдено подозрительных номеров: {len(bad_rows)}"
                )
            else:
                self.warning_box.setPlainText("Расхождений не найдено.")



        except Exception as e:
            print(e)
            traceback.print_exc()

    def save_part_excel(self):
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Сохранить испытание",
            f"Намыв.xlsx",
            "Excel Files (*.xlsx *.xls)"
        )
        if not file_path:
            return

        self.df_partii.to_excel(file_path)



    def get_namivs(self):
        try:
            item = self.list_widget2.currentItem()
            if item:
                # Получаем данные из Qt.UserRole (роль 0x100)
                db_id = item.data(Qt.UserRole)
                text = item.text()

            df = self.orkestr_db.db_show.get_namivs(db_id)

            result = vigruzka_namiv(df)

            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Сохранить испытание",
                f"Намыв_{text}.xlsx",
                "Excel Files (*.xlsx *.xls)"
            )
            if not file_path:
                return

            result.to_excel(file_path)

            wb = load_workbook(file_path)
            ws = wb["Sheet1"]

            # объединять (по номеру колонки в Excel)
            # Например, Какие столбцы  столбцы A и C
            cols_to_merge = ["B", "D", "H", "I", "J", "K", "L", "M", "N"]

            # Определяем последнюю строку
            max_row = ws.max_row

            # Объединяем по две строки: (1,2), (3,4), ...
            for col in cols_to_merge:
                row = 2  # 1-я строка — заголовки, начинаем со 2-й
                while row <= max_row:
                    # объединяем только если есть пара строк row и row+1
                    if row + 1 <= max_row:
                        ws.merge_cells(start_row=row, start_column=ws[col + str(row)].column,
                                       end_row=row + 1, end_column=ws[col + str(row)].column)
                    row += 2

            wb.save(file_path)
        except Exception as e:
            print(e)
            traceback.print_exc()


    # --- Синхронизация файлов грансоставов ---

    def sync_files_incremental(self):
        """Инкрементальное обновление: проверяет изменение файлов и добавляет только новые данные."""
        try:
            result = gran_sync.sync_all_files(self.orkestr_db.db, full_rescan=False)
            self.warning_box.setPlainText(result)
        except Exception as e:
            traceback.print_exc()
            QMessageBox.critical(self, "Ошибка", f"Ошибка синхронизации: {e}")

    def sync_files_full(self):
        """Полный перепарс всех файлов с перезаписью статусов."""
        reply = QMessageBox.question(
            self,
            "Подтверждение",
            "Будут полностью перезаписаны статусы грансоставов для всех партий.\nПродолжить?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return
        try:
            result = gran_sync.sync_all_files(self.orkestr_db.db, full_rescan=True)
            self.warning_box.setPlainText(result)
        except Exception as e:
            traceback.print_exc()
            QMessageBox.critical(self, "Ошибка", f"Ошибка перепарса: {e}")

    def open_gran_report(self):
        """Открывает окно отчёта грансоставов."""
        try:
            dialog = GranReportDialog(self.orkestr_db.db.db_file, self)
            dialog.exec_()
        except Exception as e:
            traceback.print_exc()
            QMessageBox.critical(self, "Ошибка", f"Не удалось открыть отчёт: {e}")

    def add_namiv(self):
        path_namiv, _ = QFileDialog.getOpenFileName(
            self, "Выберите файл намыва", "", "Excel Files (*.xlsx *.xls)")
        if not path_namiv:
            return

        path_tarirovki = "excelki/tarirovki.xlsx"

        try:
            df = zagr_file(path_namiv)
            df_tarirovk = zagr_tarirovki(path_tarirovki)
            udelka = 2.7

            df_unique, df_duplicates  = obrabotka_df_posle_zagr(df)

            if len(df_duplicates) > 0:

                df_itog_duplicates, spisok_otrizat_grani = rashet_gran(df_duplicates, df_tarirovk, udelka)

                if spisok_otrizat_grani:
                    QMessageBox.warning(self, "Внимание", f"Есть отрицательные граны {spisok_otrizat_grani}")

                df_itog_duplicates = df_itog_duplicates.rename(columns=config.cols_bd_rename)
                df_rashet_dubli = df_duplicates[config.cols_bd_rashet]

                self.orkestr_db.db_add.add_gran_bd(df_itog_duplicates)
                self.orkestr_db.db_add.add_gran_rashet_bd(df_rashet_dubli)

            df_itog_unique, spisok_otrizat_grani = rashet_gran(df_unique, df_tarirovk, udelka)

            if spisok_otrizat_grani:
                QMessageBox.warning(self, "Внимание", f"Есть отрицательные граны {spisok_otrizat_grani}")

            df_itog_unique = df_itog_unique.rename(columns=config.cols_bd_rename)
            df_rashet = df_unique[config.cols_bd_rashet]

            self.orkestr_db.db_add.add_gran_bd(df_itog_unique)
            self.orkestr_db.db_add.add_gran_rashet_bd(df_rashet)

            QMessageBox.information(self, "Успех",
                                    f"Граны ({len(df_unique)} шт.) в базе данных\n"
                                    f"Дубли {len(df_duplicates)}")

        except Exception as e:
            print(e)
            QMessageBox.critical(self, "Ошибка", f"Рабочий журнал не получилось добавить: {e}")
            traceback.print_exc()

    def on_item_clicked2(self, item):
        try:
            object_name = self.current_object_name
            partiya = item.text()
            db_id = item.data(Qt.UserRole)

            self.df_partii = self.orkestr_db.db_show.poln_info_part(db_id)

            print(self.df_partii)

            count = len(self.df_partii)
            count_pust_gran = len(self.df_partii[self.df_partii[config.cols_bd_rename.values()].isna().any(axis=1)])
            count_pust_wlashn = len(self.df_partii[self.df_partii['wlashn'].isna()])

            mask = self.df_partii['plotn'].isna() & self.df_partii['ukol'].notna()
            count_pust_plotn = mask.sum()

            mask_org = self.df_partii['organika'].notna()
            count_org = mask_org.sum()

            mask_udelka = self.df_partii['udelka'].notna()
            count_udelka = mask_udelka.sum()

            self.info_label.setText(f"Выбрано: {object_name}, Партия: {partiya}\n\n"
                                    f"Количество проб {count}\n"
                                    f"Количество не хватающих гранов {count_pust_gran}\n"
                                    f"Количество не хватающих влажностей {count_pust_wlashn}\n"
                                    f"Количество не хватающих плотностей {count_pust_plotn}\n\n"
                                    f"Количество сделанной органики {count_org}\n"
                                    f"Количество сделанной уделки {count_udelka}")
            self.warning_box.clear()
        except Exception as e:
            print(e)
            traceback.print_exc()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    app.setStyleSheet(qdarkstyle.load_stylesheet_pyqt5())
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
    print('Jello')
